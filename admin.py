import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

# Function to create or load Excel workbook and worksheets
def setup_excel():
    try:
        # Load existing workbook if it exists
        wb = openpyxl.load_workbook('attendance.xlsx')
    except FileNotFoundError:
        # Create a new workbook if it doesn't exist
        wb = Workbook()
        wb.save('attendance.xlsx')
    
    # Create or load worksheets for users and attendance
    if 'users' not in wb.sheetnames:
        wb.create_sheet('users')
        users_ws = wb['users']
        users_ws.append(['Username', 'Password', 'Role'])
    else:
        users_ws = wb['users']
        
    if 'attendance' not in wb.sheetnames:
        wb.create_sheet('attendance')
        attendance_ws = wb['attendance']
        attendance_ws.append(['Username', 'Subject', 'Date'])
    else:
        attendance_ws = wb['attendance']
    
    return wb, users_ws, attendance_ws

# Function to add a new user
def add_user(username, password, role, users_ws):
    users_ws.append([username, password, role])

# Function to check credentials
def authenticate(username, password, users_ws):
    for row in users_ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return row
    return None

# Function to mark attendance
def mark_attendance(username, subject, attendance_ws):
    date = datetime.now().strftime("%Y-%m-%d")
    attendance_ws.append([username, subject, date])

# Function to retrieve attendance
def get_attendance(attendance_ws):
    return attendance_ws.values

# Streamlit UI and other parts of the code remain the same
